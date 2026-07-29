import csv
import io
import json
import math

from PIL import Image, ImageDraw
from openpyxl import Workbook


def load_categories(json_file):
    data = json.load(json_file)

    if "categories" not in data:
        raise ValueError("notes.json 中找不到 categories")

    categories = {}

    for item in data["categories"]:
        if "id" not in item or "name" not in item:
            raise ValueError(
                "每個 category 都必須有 id 和 name"
            )

        class_id = int(item["id"])
        class_name = str(item["name"]).strip()

        if not class_name:
            raise ValueError(
                f"class_id {class_id} 的名稱是空白"
            )

        if class_id in categories:
            raise ValueError(
                f"重複的 class_id：{class_id}"
            )

        categories[class_id] = class_name

    return categories


def load_yolo_txt(txt_file):
    boxes = []

    for line_number, raw_line in enumerate(
        txt_file,
        start=1,
    ):
        line = raw_line.strip()

        if not line:
            continue

        parts = line.split()

        if len(parts) != 5:
            raise ValueError(
                f"第 {line_number} 列不是五欄：{line}"
            )

        try:
            class_id = int(parts[0])
            x = float(parts[1])
            y = float(parts[2])
            w = float(parts[3])
            h = float(parts[4])
        except ValueError as error:
            raise ValueError(
                f"第 {line_number} 列包含無效數字：{line}"
            ) from error

        boxes.append({
            "line_number": line_number,
            "class_id": class_id,
            "x": x,
            "y": y,
            "w": w,
            "h": h,
        })

    if not boxes:
        raise ValueError("TXT 中沒有 bounding boxes")

    return boxes


def validate_box(box, categories):
    errors = []
    warnings = []

    class_id = box["class_id"]
    x = box["x"]
    y = box["y"]
    w = box["w"]
    h = box["h"]

    if class_id not in categories:
        errors.append(
            f"未知 class_id：{class_id}"
        )

    finite_fields = {}

    for field_name, value in [
        ("x", x),
        ("y", y),
        ("w", w),
        ("h", h),
    ]:
        finite_fields[field_name] = math.isfinite(value)

        if not finite_fields[field_name]:
            errors.append(
                f"{field_name} 必須是有限數值：{value}"
            )
            continue

        if value < 0 or value > 1:
            errors.append(
                f"{field_name} 不在 0–1：{value}"
            )

    if finite_fields["w"] and w <= 0:
        errors.append("w 必須大於 0")

    if finite_fields["h"] and h <= 0:
        errors.append("h 必須大於 0")

    if all(finite_fields.values()):
        left = x - w / 2
        top = y - h / 2
        right = x + w / 2
        bottom = y + h / 2

        if left < 0:
            errors.append("框的左邊超出圖片")

        if top < 0:
            errors.append("框的上方超出圖片")

        if right > 1:
            errors.append("框的右邊超出圖片")

        if bottom > 1:
            errors.append("框的下方超出圖片")

    if (
        finite_fields["w"]
        and finite_fields["h"]
        and (w > 0.5 or h > 0.5)
    ):
        warnings.append(
            "此 bounding box 非常大，請人工確認"
        )

    return errors, warnings


def yolo_to_pixels(
    box,
    image_width,
    image_height,
):
    x = box["x"]
    y = box["y"]
    w = box["w"]
    h = box["h"]

    left = (x - w / 2) * image_width
    top = (y - h / 2) * image_height
    right = (x + w / 2) * image_width
    bottom = (y + h / 2) * image_height

    return {
        "left": round(left),
        "top": round(top),
        "right": round(right),
        "bottom": round(bottom),
        "width": round(right - left),
        "height": round(bottom - top),
    }


def draw_boxes(
    image,
    boxes,
    categories,
    show_labels=True,
):
    output = image.convert("RGB").copy()
    draw = ImageDraw.Draw(output)

    image_width, image_height = output.size

    for box in boxes:
        errors, warnings = validate_box(
            box,
            categories,
        )

        coordinates_are_drawable = (
            all(
                math.isfinite(box[field_name])
                for field_name in ("x", "y", "w", "h")
            )
            and box["w"] > 0
            and box["h"] > 0
        )

        if not coordinates_are_drawable:
            continue

        pixel_box = yolo_to_pixels(
            box,
            image_width,
            image_height,
        )

        class_id = box["class_id"]
        class_name = categories.get(
            class_id,
            "UNKNOWN",
        )

        if errors:
            box_color = "red"
        elif warnings:
            box_color = "darkorange"
        else:
            box_color = "blue"

        coordinates = [
            pixel_box["left"],
            pixel_box["top"],
            pixel_box["right"],
            pixel_box["bottom"],
        ]

        draw.rectangle(
            coordinates,
            outline=box_color,
            width=1,
        )

        if show_labels:
            label = f"{class_id} · {class_name}"

            draw.text(
                (
                    pixel_box["left"],
                    max(0, pixel_box["top"] - 14),
                ),
                label,
                fill=box_color,
            )

    return output


def make_csv(boxes):
    output = io.StringIO(newline="")

    writer = csv.writer(output)

    writer.writerow([
        "class_id",
        "x",
        "y",
        "w",
        "h",
    ])

    for box in boxes:
        writer.writerow([
            box["class_id"],
            f'{box["x"]:.6f}',
            f'{box["y"]:.6f}',
            f'{box["w"]:.6f}',
            f'{box["h"]:.6f}',
        ])

    return output.getvalue()


def sonata_name_from_page(page_name):
    stem = str(page_name)

    if stem.lower().endswith(".txt"):
        stem = stem[:-4]

    name_parts = stem.rsplit("-", 2)

    if (
        len(name_parts) != 3
        or not name_parts[1].isdigit()
        or not name_parts[2].isdigit()
    ):
        raise ValueError(
            f"無法辨識頁面檔名格式：{page_name}"
        )

    return name_parts[0]


def _safe_worksheet_title(page_name, used_titles):
    invalid_characters = set("[]:*?/\\")
    cleaned_name = "".join(
        "_"
        if character in invalid_characters
        else character
        for character in str(page_name)
    ).strip().strip("'")

    if not cleaned_name:
        cleaned_name = "Page"

    if cleaned_name.startswith(("=", "+", "-", "@")):
        cleaned_name = f"Page_{cleaned_name}"

    base_title = cleaned_name[:31]
    candidate = base_title
    suffix_number = 2

    while candidate.casefold() in used_titles:
        suffix = f"_{suffix_number}"
        candidate = (
            f"{base_title[:31 - len(suffix)]}{suffix}"
        )
        suffix_number += 1

    used_titles.add(candidate.casefold())
    return candidate


def make_xlsx(page_boxes):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.append([
        "page_name",
        "box_count",
    ])

    used_titles = {"summary"}

    for page_name in sorted(page_boxes):
        boxes = page_boxes[page_name]
        summary_row = summary_sheet.max_row + 1
        page_name_cell = summary_sheet.cell(
            row=summary_row,
            column=1,
        )
        page_name_cell.value = str(page_name)
        page_name_cell.data_type = "s"
        summary_sheet.cell(
            row=summary_row,
            column=2,
            value=len(boxes),
        )

        sheet_title = _safe_worksheet_title(
            page_name,
            used_titles,
        )

        sheet = workbook.create_sheet(
            title=sheet_title,
        )

        sheet.append([
            "class_id",
            "x",
            "y",
            "w",
            "h",
        ])

        for box in boxes:
            sheet.append([
                box["class_id"],
                box["x"],
                box["y"],
                box["w"],
                box["h"],
            ])

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:E{sheet.max_row}"
        )

        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 14

        for row in sheet.iter_rows(
            min_row=2,
            min_col=2,
            max_col=5,
        ):
            for cell in row:
                cell.number_format = "0.000000"

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 12
    summary_sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()

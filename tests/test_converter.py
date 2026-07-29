import io
import math

import pytest
from openpyxl import load_workbook
from PIL import Image

from converter import (
    draw_boxes,
    load_yolo_txt,
    make_csv,
    make_xlsx,
    sonata_name_from_page,
    validate_box,
    yolo_to_pixels,
)


CATEGORIES = {
    18: "dynamicF",
    20: "dynamicP",
}


def make_box(**changes):
    box = {
        "line_number": 1,
        "class_id": 18,
        "x": 0.5,
        "y": 0.5,
        "w": 0.1,
        "h": 0.1,
    }
    box.update(changes)
    return box


def test_first_reference_box():
    box = {
        "x": 0.201429,
        "y": 0.228247,
        "w": 0.022286,
        "h": 0.017574,
    }

    result = yolo_to_pixels(
        box,
        image_width=1750,
        image_height=2333,
    )

    assert result["left"] == 333
    assert result["top"] == 512
    assert result["width"] == 39
    assert result["height"] == 41


def test_load_yolo_txt_reads_five_columns():
    boxes = load_yolo_txt(
        io.StringIO(
            "18 0.201429 0.228247 0.022286 0.017574\n"
        )
    )

    assert len(boxes) == 1
    assert boxes[0]["class_id"] == 18
    assert boxes[0]["line_number"] == 1


@pytest.mark.parametrize(
    "contents, expected_message",
    [
        ("", "沒有 bounding boxes"),
        ("18 0.1 0.2 0.3\n", "不是五欄"),
        ("18 x 0.2 0.3 0.4\n", "包含無效數字"),
    ],
)
def test_load_yolo_txt_rejects_invalid_input(
    contents,
    expected_message,
):
    with pytest.raises(
        ValueError,
        match=expected_message,
    ):
        load_yolo_txt(io.StringIO(contents))


@pytest.mark.parametrize(
    "field_name",
    ["x", "y", "w", "h"],
)
def test_validate_box_rejects_non_finite_values(
    field_name,
):
    box = make_box()
    box[field_name] = math.nan

    errors, warnings = validate_box(
        box,
        CATEGORIES,
    )

    assert any(
        "必須是有限數值" in message
        for message in errors
    )
    assert warnings == []


def test_validate_box_rejects_unknown_class_and_bounds():
    box = make_box(
        class_id=999,
        x=0.01,
        w=0.1,
    )

    errors, _ = validate_box(
        box,
        CATEGORIES,
    )

    assert any(
        "未知 class_id" in message
        for message in errors
    )
    assert "框的左邊超出圖片" in errors


def test_draw_boxes_skips_non_finite_coordinates():
    image = Image.new(
        "RGB",
        (100, 100),
        "white",
    )

    result = draw_boxes(
        image,
        [make_box(x=math.nan)],
        CATEGORIES,
    )

    assert result.getpixel((50, 50)) == (
        255,
        255,
        255,
    )


def test_make_csv_has_five_yolo_columns():
    csv_data = make_csv([make_box()])
    rows = csv_data.strip().splitlines()

    assert rows[0] == "class_id,x,y,w,h"
    assert rows[1] == (
        "18,0.500000,0.500000,0.100000,0.100000"
    )


def test_sonata_name_accepts_hyphenated_prefix():
    assert sonata_name_from_page(
        "Composer-Name_Op090-01-06.txt"
    ) == "Composer-Name_Op090"


@pytest.mark.parametrize(
    "page_name",
    [
        "Beethoven_Op090",
        "Beethoven_Op090-first-01",
        "Beethoven_Op090-01-last",
    ],
)
def test_sonata_name_rejects_invalid_page_name(
    page_name,
):
    with pytest.raises(
        ValueError,
        match="無法辨識頁面檔名格式",
    ):
        sonata_name_from_page(page_name)


def test_make_xlsx_sanitizes_titles_and_formula_text():
    workbook_data = make_xlsx({
        "=1+1": [make_box()],
        "bad/name": [make_box()],
    })
    workbook = load_workbook(
        io.BytesIO(workbook_data),
        data_only=False,
    )

    summary_name_cell = workbook["Summary"]["A2"]

    assert summary_name_cell.value == "=1+1"
    assert summary_name_cell.data_type == "s"
    assert "Page_=1+1" in workbook.sheetnames
    assert "bad_name" in workbook.sheetnames


def test_make_xlsx_creates_unique_31_character_titles():
    shared_prefix = "a" * 31
    workbook_data = make_xlsx({
        f"{shared_prefix}1": [make_box()],
        f"{shared_prefix}2": [make_box()],
    })
    workbook = load_workbook(
        io.BytesIO(workbook_data)
    )
    page_sheets = [
        name
        for name in workbook.sheetnames
        if name != "Summary"
    ]

    assert len(page_sheets) == 2
    assert len(set(page_sheets)) == 2
    assert all(
        len(name) <= 31
        for name in page_sheets
    )
